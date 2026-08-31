"""把用户上传的参考文件读成给模型看的文本：表格摊平，纯文本原样截取。

⚠ **不落库、不存对象存储**。读完就把内容交给模型，会话里存的是那段文本。
存文件要连带一整套生命周期（谁能读、什么时候删、删了引用怎么办），而这份
文件的用处只有一次——它是这一轮对话的参考资料，不是资产。

⚠ 内容都有上限，且**截断了要说出来**。一张几千行的表或一整本手册塞进上下文，
会把技能正文与工作面快照一起挤掉，而挤掉了哪一段从外面完全看不出来。

⚠ CSV 用 `utf-8-sig` 读：现场的点表十有八九是从组态软件导出、再用 Excel 存过
一道的，带 BOM 是常态。不剥的话首列表头会多一个看不见的字符，而「为什么第一列
匹配不上」查起来极费劲（前端那份 CSV 导入踩过同一个坑）。
"""

import csv
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook

# 行列上限。列多是横表（一行一个设备），行多是竖表（一行一个点位），两种都有
MAX_ROWS = 200
MAX_COLUMNS = 30
# 单元格文本上限：整段说明塞进一格是常事，而它对匹配没有帮助
MAX_CELL = 120
# 纯文本的字符上限，量级与截断后的表持平：附件是参考资料，不是全文检索库
MAX_TEXT_CHARS = 24_000

_XLSX_SUFFIXES = (".xlsx", ".xlsm")
# 摊成表头加数据行的那几种。⚠ 与解码器共用同一份：各写一份的表现是
# 「选得中的文件传上去被拒」，而两边看起来都对
TABLE_SUFFIXES = (*_XLSX_SUFFIXES, ".csv")
# 逐字读的纯文本类。⚠ 名单是显式的：兜底「读得动就收」会把 pdf / 图片这类
# 二进制也放进来，解出一堆乱码还占满上下文
TEXT_SUFFIXES = (
    ".txt",
    ".md",
    ".markdown",
    ".json",
    ".log",
    ".yaml",
    ".yml",
    ".xml",
    ".ini",
    ".toml",
)


class UnsupportedTable(ValueError):
    """认不出的文件类型。"""


@dataclass(frozen=True)
class ParsedTable:
    """读出来的一份内容：表格走 columns/rows，纯文本走 body，二选一。"""

    columns: list[str] = field(default_factory=list[str])
    rows: list[list[str]] = field(default_factory=list[list[str]])
    is_truncated: bool = False
    # 原始行数（截断前）。⚠ 截断了必须让人看得见到底少了多少
    total_rows: int = 0
    # 纯文本文件的正文；表格文件恒为空串
    body: str = ""


def parse_table(filename: str, content: bytes) -> ParsedTable:
    """按文件名后缀选解析方式。

    Args: filename, content。
    """
    lowered = filename.lower()
    if lowered.endswith(_XLSX_SUFFIXES):
        return _shape(_read_xlsx(content))
    if lowered.endswith(".csv"):
        return _shape(_read_csv(content))
    if lowered.endswith(TEXT_SUFFIXES):
        return _shape_text(_decode(content))
    raise UnsupportedTable(
        "只认得表格（.xlsx / .xlsm / .csv）或纯文本（.txt / .md / .json 等）"
    )


def _read_csv(content: bytes) -> list[list[str]]:
    """读 CSV。⚠ `utf-8-sig` 剥 BOM；解码不了就退回 GBK——组态软件导出的
    点表常是本地编码，而那时报「不是合法 UTF-8」对用户毫无帮助。
    """
    text = _decode(content)
    return [list(row) for row in csv.reader(io.StringIO(text))]


def _decode(content: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise UnsupportedTable("这个文件的编码认不出来")


def _read_xlsx(content: bytes) -> list[list[str]]:
    """读 xlsx 的第一张工作表。

    ⚠ `data_only=True`：取公式**算出来的值**而不是公式本身。取到 `=A1*2`
    这种字符串的话，匹配永远对不上，而表格在 Excel 里看着一切正常。
    """
    book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = book.worksheets[0] if book.worksheets else None
        if sheet is None:
            return []
        return [
            [_cell(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        book.close()


def _cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()[:MAX_CELL]


def _shape(grid: list[list[str]]) -> ParsedTable:
    """把二维格摊成表头 + 数据行，顺带截断。

    ⚠ 整行都是空的直接丢：导出的表尾常挂着几十行空格，它们进了上下文就是纯噪音。

    Args: grid。
    """
    live = [row for row in grid if any(cell.strip() for cell in row)]
    if not live:
        return ParsedTable()
    header = [cell.strip() for cell in live[0][:MAX_COLUMNS]]
    body = live[1:]
    kept = [
        [cell.strip() for cell in row[:MAX_COLUMNS]] for row in body[:MAX_ROWS]
    ]
    return ParsedTable(
        columns=header,
        rows=kept,
        is_truncated=len(body) > MAX_ROWS,
        total_rows=len(body),
    )


def _shape_text(text: str) -> ParsedTable:
    """纯文本按字符截断。行数口径是全文行数，截断了从数字上看得出来。

    Args: text。
    """
    stripped = text.strip()
    if stripped == "":
        return ParsedTable()
    return ParsedTable(
        body=stripped[:MAX_TEXT_CHARS],
        is_truncated=len(stripped) > MAX_TEXT_CHARS,
        total_rows=len(stripped.splitlines()),
    )


def to_text(table: ParsedTable) -> str:
    """摊成一段给模型看的文本。

    ⚠ 表格用竖线分隔而不是 JSON：同样的内容，表格形态的 token 数少得多，
    而模型对「一行是一条」这件事看得更准。纯文本原样给。

    Args: table。
    """
    if table.body != "":
        if table.is_truncated:
            note = f"（只截了开头一段，全文共 {table.total_rows} 行）"
            return f"{table.body}\n{note}"
        return table.body
    if not table.columns:
        return "（这个文件里没有读到任何内容）"
    lines = [" | ".join(table.columns)]
    lines.extend(" | ".join(row) for row in table.rows)
    if table.is_truncated:
        lines.append(
            f"（只列了前 {MAX_ROWS} 行，整张表共 {table.total_rows} 行）"
        )
    return "\n".join(lines)


def summary_of(table: ParsedTable) -> str:
    """给人看的一句概况：这份文件有多大、截没截。

    ⚠ 算在服务端而不是界面上：界面各算一份的话，「几列几行」的口径会与
    真正喂给模型的那一份漂开，而用户看到的是前者。

    Args: table。
    """
    if table.body != "":
        shape = f"{table.total_rows} 行"
    elif table.columns:
        shape = f"{len(table.columns)} 列 × {table.total_rows} 行"
    else:
        return "空文件"
    return f"{shape}，已截断" if table.is_truncated else shape
